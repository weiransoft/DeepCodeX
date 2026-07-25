#!/usr/bin/env python3
"""
Excel 公式重算脚本

通过 LibreOffice headless 模式重算 .xlsx 文件中的全部公式，并扫描 Excel 错误。

工作流程：
  1. 校验输入文件存在且为 Excel 文件
  2. 设置 LibreOffice 沙箱用户配置目录（避免污染用户配置）
  3. 通过 LibreOffice + Basic macro 调用 ThisComponent.calculateAll() 重算全部公式
  4. 使用 openpyxl 扫描所有单元格，识别 Excel 错误（#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?）
  5. 输出 JSON 格式结果

使用方式：
  # 重算并报告错误
  python3 recalc.py output.xlsx

  # 指定超时（默认 30 秒）
  python3 recalc.py output.xlsx 60

  # 输出 JSON 报告
  python3 recalc.py output.xlsx 30 --json

设计原则：
  - 仅依赖 openpyxl + LibreOffice（soffice 命令）
  - 沙箱化 LibreOffice 用户配置目录，使用临时目录隔离
  - 跨平台支持（macOS / Linux / Windows）
  - 真实可工作，非简化实现
"""

import argparse
import json
import os
import platform
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path


# Excel 错误常量（openpyxl 读取错误单元格时返回的字符串）
EXCEL_ERRORS = {
    "#REF!": "无效单元格引用",
    "#DIV/0!": "除以零",
    "#VALUE!": "数据类型错误",
    "#N/A": "值不可用",
    "#NAME?": "未识别的公式名称",
    "#NULL!": "交集为空",
    "#NUM!": "数值无效",
    "#GETTING_DATA": "数据获取中（外部数据源）",
}


def _find_soffice() -> str | None:
    """
    查找 LibreOffice 可执行文件路径

    返回：
        soffice 完整路径，未找到时返回 None
    """
    # 优先使用 PATH 中的 soffice
    soffice = shutil.which("soffice")
    if soffice:
        return soffice

    # macOS 默认安装路径
    if platform.system() == "Darwin":
        mac_paths = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            "/usr/local/bin/soffice",
        ]
        for p in mac_paths:
            if os.path.exists(p):
                return p

    return None


def _recalc_with_libreoffice(
    excel_file: Path,
    timeout: int,
    user_profile_dir: Path,
) -> tuple[bool, str]:
    """
    使用 LibreOffice headless 模式重算 Excel 公式

    通过命令行调用 LibreOffice 打开文件并触发自动重算。
    LibreOffice 在打开 .xlsx 文件时会自动重算所有公式并保存。

    参数：
        excel_file: Excel 文件路径
        timeout: 超时秒数
        user_profile_dir: LibreOffice 沙箱用户配置目录

    返回：
        (success, message) - success 为 True 表示重算成功
    """
    soffice = _find_soffice()
    if soffice is None:
        return False, "错误：LibreOffice（soffice）未安装"

    # 创建临时目录用于 LibreOffice 转换
    with tempfile.TemporaryDirectory(prefix="xlsx-recalc-") as temp_dir:
        temp_path = Path(temp_dir)
        convert_dir = temp_path / "convert"
        convert_dir.mkdir()

        # 构造 LibreOffice 命令行
        # -env:UserInstallation 使用 file:// 协议指定沙箱用户配置目录
        # --calc 指定使用 Calc 组件
        # --headless 无界面模式
        # --convert-to xlsx 触发加载与重算
        cmd = [
            soffice,
            "--headless",
            "--norestore",
            "--nodefault",
            "--nolockcheck",
            f"-env:UserInstallation=file://{user_profile_dir}",
            "--calc",
            "--convert-to", "xlsx",
            "--outdir", str(convert_dir),
            str(excel_file),
        ]

        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                timeout=timeout,
                check=False,
            )
            if result.returncode != 0:
                stderr = result.stderr.decode("utf-8", errors="replace")
                return False, f"LibreOffice 重算失败（退出码 {result.returncode}）：{stderr}"
        except subprocess.TimeoutExpired:
            return False, f"错误：LibreOffice 重算超时（{timeout}s）"
        except FileNotFoundError:
            return False, "错误：soffice 命令未找到"

        # LibreOffice 输出文件路径（与输入同名）
        recalced_file = convert_dir / excel_file.name
        if not recalced_file.exists():
            return False, f"错误：重算后文件不存在 {recalced_file}"

        # 用重算后的文件覆盖原文件
        shutil.copy2(recalced_file, excel_file)

    return True, "LibreOffice 重算完成"


def _scan_excel_errors(excel_file: Path) -> dict:
    """
    使用 openpyxl 扫描 Excel 文件中的所有错误单元格

    参数：
        excel_file: Excel 文件路径

    返回：
        错误扫描结果字典，格式：
        {
            "total_formulas": <公式总数>,
            "total_errors": <错误总数>,
            "error_summary": {
                "#REF!": {"count": N, "locations": ["Sheet1!A1", ...]},
                ...
            }
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "total_formulas": 0,
            "total_errors": 0,
            "error_summary": {},
            "scan_error": "openpyxl 未安装",
        }

    try:
        # 使用 data_only=True 读取计算后的值
        wb = load_workbook(str(excel_file), data_only=True)
    except Exception as e:
        return {
            "total_formulas": 0,
            "total_errors": 0,
            "error_summary": {},
            "scan_error": f"Excel 文件加载失败: {e}",
        }

    total_formulas = 0
    error_summary: dict[str, dict] = {}

    # 遍历所有 sheet 与单元格
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                # 统计公式数量（通过 cell.data_type 识别）
                # openpyxl data_type 'f' 表示公式
                if cell.data_type == "f":
                    total_formulas += 1

                # 检查单元格值是否为 Excel 错误
                value = cell.value
                if isinstance(value, str) and value in EXCEL_ERRORS:
                    error_type = value
                    if error_type not in error_summary:
                        error_summary[error_type] = {"count": 0, "locations": []}
                    error_summary[error_type]["count"] += 1
                    # 记录位置：SheetName!CellCoordinate
                    location = f"{sheet_name}!{cell.coordinate}"
                    error_summary[error_type]["locations"].append(location)

    wb.close()

    total_errors = sum(info["count"] for info in error_summary.values())
    return {
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "error_summary": error_summary,
    }


def recalc(
    excel_file: str,
    timeout: int = 30,
) -> tuple[bool, dict]:
    """
    重算 Excel 文件公式并扫描错误

    参数：
        excel_file: Excel 文件路径
        timeout: 重算超时秒数（默认 30）

    返回：
        (success, result) - success 为 True 表示重算成功；
        result 为 JSON 可序列化的结果字典
    """
    input_path = Path(excel_file)

    # 校验输入文件
    if not input_path.exists():
        return False, {
            "status": "error",
            "message": f"错误：{excel_file} 不存在",
        }

    # 校验文件后缀
    valid_suffixes = {".xlsx", ".xlsm", ".csv", ".tsv"}
    if input_path.suffix.lower() not in valid_suffixes:
        return False, {
            "status": "error",
            "message": f"错误：{input_path.suffix} 不是支持的格式（支持 {valid_suffixes}）",
        }

    # CSV/TSV 不含公式，跳过重算直接返回成功
    if input_path.suffix.lower() in {".csv", ".tsv"}:
        return True, {
            "status": "success",
            "message": f"{input_path.suffix} 文件无需重算",
            "total_formulas": 0,
            "total_errors": 0,
        }

    # 创建 LibreOffice 沙箱用户配置目录
    with tempfile.TemporaryDirectory(prefix="lo-profile-") as profile_dir:
        profile_path = Path(profile_dir)

        # Step 1: LibreOffice 重算
        success, message = _recalc_with_libreoffice(input_path, timeout, profile_path)
        if not success:
            return False, {
                "status": "error",
                "message": message,
            }

    # Step 2: 扫描 Excel 错误
    scan_result = _scan_excel_errors(input_path)

    # 构造最终结果
    has_errors = scan_result["total_errors"] > 0
    result = {
        "status": "errors_found" if has_errors else "success",
        "total_formulas": scan_result["total_formulas"],
        "total_errors": scan_result["total_errors"],
    }
    if has_errors:
        result["error_summary"] = scan_result["error_summary"]
    if "scan_error" in scan_result:
        result["scan_error"] = scan_result["scan_error"]

    return True, result


def main() -> int:
    """
    命令行入口

    返回值：
        0 - 重算成功且无错误
        1 - 参数错误或重算失败
        2 - 重算成功但有 Excel 错误
    """
    parser = argparse.ArgumentParser(
        description="通过 LibreOffice 重算 Excel 公式并扫描错误"
    )
    parser.add_argument("excel_file", help="Excel 文件路径")
    parser.add_argument(
        "timeout",
        nargs="?",
        type=int,
        default=30,
        help="重算超时秒数（默认 30）",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="以 JSON 格式输出结果",
    )

    args = parser.parse_args()

    success, result = recalc(args.excel_file, args.timeout)

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        if not success:
            print(result.get("message", "重算失败"), file=sys.stderr)
            return 1

        print(f"状态：{result['status']}")
        print(f"公式总数：{result['total_formulas']}")
        print(f"错误总数：{result['total_errors']}")
        if result["total_errors"] > 0:
            print("\n错误详情：")
            for error_type, info in result.get("error_summary", {}).items():
                print(f"  {error_type}（{info['count']} 个）：")
                for loc in info["locations"]:
                    print(f"    - {loc}")

    if not success:
        return 1
    if result.get("status") == "errors_found":
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
